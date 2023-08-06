"""
Usage:
    generate_bash.py [--input_file=FILE] [--local_app_path=STR] [--local_vsfs_path=STR]  [--run_hard=INT] [--output=STR]
    generate_bash.py --version
    generate_bash.py -h | --help

Options:
 -h --help
 --input_file FILE      A path to the input database containg command lines. It's inside this package, testVSFS folder,
                        filename is Public apps by sevenbridges.
 --local_app_path STR   This is a path to the folder containing all the bix demo apps. Also, inside the testVSFS.
 --local_vsfs_path STR  Path to the local point where vsfs is mounted. Example: /home/mladenl/vsfs/
 --run_hard INT         Do you want to run computationally intensive apps also. Default mode is no. Put non-zero int if
                        you want to run them.
 --output STR           A path to the output file. By defualt it will be in testVSFS folder. Needs to have .xlsx extension
"""

from openpyxl import load_workbook
from docopt import docopt
import subprocess
import json
import os
from collections import defaultdict


def main():
    args = docopt(__doc__, version='0.1.0')
    filepath = args["--input_file"]
    local_vsfs_path = args["--local_vsfs_path"]
    run_hard = args["--run_hard"]
    local_app_path = args["--local_app_path"]
    output = args["--output"]

    # query formation
    wb = load_workbook(filename=filepath)
    sheet_ranges = wb['Tools'] # select sheet from wb
    base_entries = range(4, 200)
    entry_cells = ['F']
    computationally_requierd = ['I']

    filename, file_extension = os.path.splitext(filepath)
    if output:
        out_wb = output
    else:
        out_wb = filename+"newout.xlsx"

    # output wb
    subprocess.call(['cp', filepath, out_wb])
    results_wb = load_workbook(filename=out_wb)
    out_sheet = results_wb['Tools']
    out_result_cell = ['M']
    out_sha_cell = ['N']
    diff_cell = ['P']
    json_cell = ['K']

    # running the comands
    for entry in base_entries:
        for cell in entry_cells:

            cmd_string = str(sheet_ranges[cell+str(entry)].value)
            cmd_string = preprocessed_cmd(cmd_string, local_vsfs_path, local_app_path)

            cmd = cmd_string.split(" ")
            intensive_app = str(sheet_ranges[computationally_requierd[0]+str(entry)].value)
            print ("cmd:", cmd)
            message, out_json = run_cmd(cmd, intensive_app, run_hard)
            try:
                out_sheet[out_result_cell[0]+str(entry)] = message
                validation_json = sheet_ranges[json_cell[0]+str(entry)].value
                if message == 'Pass' and validation_json and out_json:
                    dif_message = compare_sizes(validation_json, out_json)

                    out_sheet[out_sha_cell[0] + str(entry)] = ' '.join(dif_message.keys())
                    out_sheet[diff_cell[0] + str(entry)] = json.dumps(dif_message.get('not equal sha1'))
            except:
                'write in table failed'
    results_wb.save(out_wb)


# this function is used to preprocess comand lines from the table (databse) in order to ajdust them for execution on the
# local machine. first argument is the command line, secon is a path t vsfs mounted directory on local machine and third
# argument is path to the folder containing json filed of demo apps.
def preprocessed_cmd(cmd, vsfs_path = 0, app_path = 0):
    if app_path:
        cmd = cmd.replace('/home/mladenl/bix-demo-apps/', app_path)
    if vsfs_path:
        cmd = cmd.replace('/home/mladenl/vsfs/', vsfs_path)
    return cmd


def run_cmd(cmd, intensive_app, run_hard):
    if intensive_app == 'y' and not run_hard:
        print 'Skipping'
        return 'Skipping', 0
    else:
        try:
            out = subprocess.check_output(cmd)
            print ("out:", out)
            return 'Pass', out
        except subprocess.CalledProcessError as e:
            print e.returncode
            print e.output
            print e.cmd
            return 'Failed CalledProcessError', e.returncode
        except OSError:
            print OSError.message
            print OSError.errno
            return'Failed OSError', OSError.message
        except:
            print('Failed', cmd)
            return 'Fail', 0


# function used for adjustment remaping paths from json to paths of files when mounted with vsfs
def process_plat_path(p_path):
    n_path = p_path.replace('/sbgenomics/', '/home/mladenl/vsfs/')
    return n_path


def compare_sizes(j_platform, j2):
    if j_platform and j2:
        json_plat = json.loads(j_platform)
        json2 = json.loads(j2)
        result = defaultdict(list)

        for key in json_plat:
            try:
                path1 = process_plat_path(json_plat[key]["path"])
                path2 = json2[key]["path"]
                sha1 = get_checksum(path1)
                sha2 = get_checksum(path2)

                if sha1 == sha2:
                    print ('equal sha')
                    result['Equal sha1'].append(0)
                else:
                    print ('diff sha'+' '+path1+' '+path2)
                    difference = compare_files(path1, path2)
                    print difference
                    result['not equal sha1'].append({path1: difference})
            except:
                print('something wrong')
                result['something wrong'].append(0)
            return result

    else:
        return {'no jsons': 0}


def compare_files(filep1, filep2):
    try:
        diff = subprocess.check_output(['diff', filep1, filep2])
        print 'no diff :/'
    except subprocess.CalledProcessError as e:
        if e.returncode == 1:
            if len(e.output) < 2000:
                print e.output
            return str(e.output)
        else:
            return 'some diff err'


def get_checksum(file_path):
    try:
        sha1_out = subprocess.check_output(['sha1sum', file_path])
        file_sha1 = sha1_out.split(' ')[0]
        return file_sha1
    except:
        print ("no-sha 1")
        return "no-sha 1"


def get_path(json_file):    # function receives a sting json
    # change this, return dict
    parsed_json = json.loads(json_file)
    paths = {}
    for key in parsed_json:
        try:
            paths[key]= parsed_json[key]["path"]
            # paths.append(parsed_json[key]["path"])
        except:
            return "path failed"
    print paths


if __name__ == '_main_':
    main()
