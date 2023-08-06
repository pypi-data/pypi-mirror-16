import openpyxl

def data_from_xl(xl_path):
    input_xl = openpyxl.load_workbook(xl_path)
    input_sheet = input_xl.get_sheet_by_name('Sheet 1')
    event_name = input_sheet['A2'].value
    event_date = input_sheet['B2'].value
    catalog_ids = []
    for i in range(2, input_sheet.max_row + 1):
        cat_id = input_sheet.cell(row=i, column=3)
        catalog_ids.append(cat_id.value)
    return event_name, event_date, catalog_ids

xl_in = raw_input("path to excel file: ")
print "User input: ", xl_in

# ingest event information through csv
info = data_from_xl(xl_in)
name = info[0]
date = info[1]
cat_ids = info[2]
print info



