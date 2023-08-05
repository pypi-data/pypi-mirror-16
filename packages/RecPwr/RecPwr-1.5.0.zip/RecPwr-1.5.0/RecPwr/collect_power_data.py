from __future__ import division, print_function  # part one of peak detect imports

import os
import subprocess
import sys
import autoit
import time
import csv
from datetime import datetime
import easygui
from easygui import msgbox
from easygui import EgStore
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Alignment
import statistics
import json

# remaining peak detect imports
from detect_peaks import detect_peaks
import numpy as np
sys.path.insert(1, r'./../functions')  # add to pythonpath


# BEGIN CREATION OF SETTINGS CLASS THAT IS USED TO REMEMBER USER INPUT ABOUT TEST RUNS
# -----------------------------------------------------------------------
# define a class named Settings as a subclass of EgStore
# -----------------------------------------------------------------------
class Settings(EgStore):

    def __init__(self, filename):  # filename is required
        # -------------------------------------------------
        # Specify default/initial values for variables that
        # this particular application wants to remember.
        # -------------------------------------------------
        self.info = {"Subject": "Shawn C",
                    "Pack Model": "PFv2.3",
                    "Pack Number [XXX]": "000",
                    "Weight [lb]": "60",
                    "Speed [MPH]": "2.0",
                    "Walk Style [Normal/LP]": "LP",
                    "ECM": "ECM2-006",
                    "Emulation Resistance": "10.0",
                    "Cap Bank": "22",
                    "Spring Rate [lb/in]": "20.0",
                    "Load Type": "Digital Load",
                    "Load [V]": "15.00",
                    "Drivetrain": "Rack and Pinion",
                    "Clutch": "5",
                    "Comments": "",
                    }

        # -------------------------------------------------
        # For subclasses of EgStore, these must be
        # the last two statements in  __init__
        # -------------------------------------------------
        self.filename = filename  # this is required
        self.restore()            # restore values from the storage file if possible

    def create_field_names(self):
        """
        Initializes the fields that will be displayed as text boxes in the prompt window
        :param self:
        :return:
        """

        field_names = ["Subject",
                        "Pack Model",
                        "Pack Number [XXX]",
                        "Weight [lb]",
                        "Speed [MPH]",
                        "Walk Style [Normal/LP]",
                        "ECM",
                        "Emulation Resistance",
                        "Cap Bank",
                        "Spring Rate [lb/in]",
                        "Load Type",
                        "Load [V]",
                        "Drivetrain",
                        "Clutch",
                        "Comments"
                     ]
        return field_names

    def create_field_values(self, field_names):
        """
        Instantiate the default values to display in the fields of the window prompt boxes
        :param self:
        :param field_names:
        :return:
        """
        field_values = []
        for x in field_names:
            field_values.append(self.info.get(x))
        return field_values

    def update_info_values(self, user_entered_value_list):
        self.info["Subject"] = user_entered_value_list[0]
        self.info["Pack Model"] = user_entered_value_list[1]
        self.info["Pack Number [XXX]"] = user_entered_value_list[2]
        self.info["Weight [lb]"] = user_entered_value_list[3]
        self.info["Speed [MPH]"] = user_entered_value_list[4]
        self.info["Walk Style [Normal/LP]"] = user_entered_value_list[5]
        self.info["ECM"] = user_entered_value_list[6]
        self.info["Emulation Resistance"] = user_entered_value_list[7]
        self.info["Cap Bank"] = user_entered_value_list[8]
        self.info["Spring Rate [lb/in]"] = user_entered_value_list[9]
        self.info["Load Type"] = user_entered_value_list[10]
        self.info["Load [V]"] = user_entered_value_list[11]
        self.info["Drivetrain"] = user_entered_value_list[12]
        self.info["Clutch"] = user_entered_value_list[13]
        self.info["Comments"] = user_entered_value_list[14]


def get_users_main_directory():
    """ If a user has pulled this project from the repo, this ensures program is run from their download directory
    :return:
    """
    # print( os.path.dirname(sys.argv[0]))
    return os.path.dirname(sys.argv[0])


def check_lp_power_data_folder_exists(user_dir):
    if not os.path.exists(os.path.abspath(os.path.join(user_dir, "LP Power Data"))):
        os.mkdir(os.path.abspath(os.path.join(user_dir, "LP Power Data")))


def check_database_file_exists(user_dir):
    # print(os.path.abspath(os.path.join(user_dir, "Experiment_Database.xlsx")))
    if not os.path.exists(os.path.abspath(os.path.join(user_dir, "Experiment_Database.xlsx"))):
        easygui.ccbox(title="WARNING", msg="The database file Experiment_Database.xlsx where power data is stored was not found in the RecPwr main program directory. In order to successfully Save data and to ensure that the data you capture is added sequentially with the correct Run number at the bottom of the database, you must Copy and Paste the most up to date Excel_Database.xlsx file from the Network drive into the folder '{0}' and Then click CONTINUE or rerun this program. Once testing is complete, make sure to copy this Local file back to the Network so that the Network is always up to date.".format(os.path.abspath(user_dir)))


def click_on_file(filename):
    try:
        os.startfile(filename)
    except AttributeError:
        subprocess.call(['open', filename])


def prompt_setup(trial_dict, user_dir):

    trial = Settings(os.path.abspath(os.path.join(user_dir, "default_settings.txt")))
    msg = "Enter the correct info to describe this trial if the default values need modification:"
    title = "Lightning Packs Power-Data Collection Setup"
    fieldNames = trial.create_field_names()
    fieldValues = []
    fieldValues = easygui.multenterbox(msg, title, fieldNames, trial.create_field_values(fieldNames))

    # ----------------------------------------------------------------------------------------
    # Update the settings class instance to have its fieldvalues be what we just entered,
    # then store to disk as new default run settings to appear next time program runs
    # ----------------------------------------------------------------------------------------

    trial.update_info_values(fieldValues)
    trial.store()
    trial_dict = trial.info

    return trial_dict


def make_pack_folder(p_number, w_speed, w_style, user_dir):
    root = os.path.abspath(os.path.join(user_dir, "LP Power Data"))
    target = root + "/LP-{0}".format(p_number)
    if not os.path.exists(target):
        print(" ")
        print("made new pack folder at: ", target)
        os.mkdir(target)
    return make_speed_folder(target, w_speed, w_style)


def make_speed_folder(s_target, w_speed, w_style):
    speed_target = os.path.join(s_target, "{0}MPH-{1} Walk".format(w_speed, w_style))
    if not os.path.exists(speed_target):
        os.mkdir(speed_target)
    else:
        return


def convert_csv_to_xlsx(csv_filename, dest_filename, user_dir):
    wb = load_workbook(os.path.join(user_dir, "csv_transform_template.xlsx"))
    ws = wb.get_sheet_by_name("Sheet1")

    f = open(csv_filename, 'r')
    reader = csv.reader(f)
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            s = cell
            try:
                s = float(s)
            except ValueError:
                pass
            ws.cell('%s%s' % (column_letter, (row_index + 1))).value = s
    wb.save(filename=dest_filename)


def take_abs_val(sheet, column_number):
    hr = sheet.max_row
    for i in range(4, hr):
        sheet.cell(row=i, column=column_number).value = abs(sheet.cell(row=i, column=column_number).value)
    col_name = sheet.cell(row=1, column=column_number).value
    sheet.cell(row=1, column=column_number).value = "abs({0})".format(col_name)


def gen_guess_frequency(walk_speed):
    # usual_correlation = {"1.5": 1.30, "1.9": 1.43, "2.5": 1.68, "3.5": 1.95, "4.0": 2.04}
    if walk_speed == "1.5":
        return 576
    if walk_speed == "1.9":
        return 524.5
    if walk_speed == "2.5":
        return 446
    if walk_speed == "3.5":
        return 385
    if walk_speed == "4.0":
        return 367.5
    else:
        return 469


def fix_xlsx(path, walk_speed):
    wb = load_workbook(filename=path)
    sheet1 = wb.active
    take_abs_val(sheet1, 2)  # Taking the abs_val ignores zeroing and low level noise errors!
    hr = sheet1.max_row

    # generate list of all power values over this particular 5sec interval excel sheet
    pwr_list = []
    for i in range(4, hr+1):
        pwr_list.append(sheet1.cell(row=i, column=2).value)

    # Find the single maximum power peak for the full 5sec interval
    max_peak = round(max(pwr_list), 4)

    # Detect the peaks that are over 80% of the max peak and are spaced a minimum 90% of the walk frequency apart.
    peaks = detect_peaks(pwr_list, mph=(0.8*max_peak), mpd=(0.9*gen_guess_frequency(walk_speed)), edge='rising', show=False)

    # Copy only the data that falls between the first and the last peaks to column 3 of the Excel sheet.
    for i in range(4, hr+1):
        if peaks[0] <= (i-4) <= peaks[-1]:
            sheet1.cell(row=i, column=3).value = sheet1.cell(row=i, column=2).value
        else:
            continue

    # Now calculate the average of this truncated Column 3 (D) to get the true Average Power for this 5sec Interval
    truncated_pwr_list = []
    for i in range(4, hr+1):
        if sheet1.cell(row=i, column=3).value is not None:
            truncated_pwr_list.append(sheet1.cell(row=i, column=3).value)
        else:
            continue
    avg_pwr = round(statistics.mean(truncated_pwr_list), 2)
    sheet1['D1'].value = "Avg. Power"
    sheet1['D2'].value = "(W)"
    # sheet1['D4'].value = "=ROUND(AVERAGE(B4:B15050),2)"
    sheet1['D4'].value = avg_pwr
    wb.save(filename=path)
    return avg_pwr


def recursively_repair_all_csv_files(save_text, user_dir, walk_speed):
    """
    Walks down subdirectories of the save path and updates all csvs to be xlsx files and computes their averages.
    :param user_dir:
    :param walk_speed:
    :param save_text:
    :return:run_avgs:
    """

    run_avgs = {}
    for (root, dirs, files) in os.walk(save_text.split(".psdata")[0]):
        for file in files:
            if file.endswith(".csv"):

                # CONVERT CSV OUTPUT OF PICOSCOPE TO XLSX TO ALLOW AVERAGE CALCULATIONS
                csv_path = os.path.realpath(os.path.join(root, file))
                dest_filename = csv_path.replace(".csv", ".xlsx")
                convert_csv_to_xlsx(csv_path, dest_filename, user_dir)

                # REPAIR XLSX DOCUMENT AND GENERATE DICTIONARY OF 5SEC SPLIT AVERAGES
                split_number = "{}".format(str(dest_filename).split('_')[-1].strip('.xlsx'))
                run_avgs[split_number] = fix_xlsx(dest_filename, walk_speed)

    #  AND THE TOTAL COMBINED AVERAGE
    run_avgs["total_avg"] = round(statistics.mean(run_avgs.values()), 2)
    # print("5sec split_avgs= ", run_avgs)
    return run_avgs


def print_average_power(fixed_xlsx):
    wb_fixed = load_workbook(filename=fixed_xlsx)
    sheet1_fixed = wb_fixed.get_sheet_by_name("Sheet1")
    magic_num = sheet1_fixed['D4'].value
    print("This run's power was {} Watts".format(magic_num))


def patch_experiments_database(trial, run_avgs, user_dir):
    wb_path = os.path.join(user_dir, "Experiment_Database.xlsx")
    wb_exp = load_workbook(filename=wb_path)
    data = wb_exp.get_sheet_by_name("data")
    hr = data.max_row
    new = hr + 1
    data['C{}'.format(new)].value = trial["Subject"]
    data['D{}'.format(new)].value = trial["Pack Model"]
    data['E{}'.format(new)].value = trial["Pack Number [XXX]"]
    data['J{}'.format(new)].value = float(trial["Weight [lb]"])
    data['K{}'.format(new)].value = float(trial["Speed [MPH]"])
    data['L{}'.format(new)].value = trial["Walk Style [Normal/LP]"]
    data['F{}'.format(new)].value = trial["ECM"]
    data['G{}'.format(new)].value = float(trial["Emulation Resistance"])
    data['H{}'.format(new)].value = float(trial["Cap Bank"])
    data['I{}'.format(new)].value = float(trial["Spring Rate [lb/in]"])
    data['S{}'.format(new)].value = trial["Load Type"]
    data['T{}'.format(new)].value = float(trial["Load [V]"])
    data['U{}'.format(new)].value = trial["Drivetrain"]
    data['V{}'.format(new)].value = float(trial["Clutch"])
    data['W{}'.format(new)].value = trial["Comments"]
    data['N{}'.format(new)].value = float(run_avgs["1"])
    data['O{}'.format(new)].value = float(run_avgs["2"])
    data['P{}'.format(new)].value = float(run_avgs["3"])
    data['Q{}'.format(new)].value = float(run_avgs["4"])
    data['R{}'.format(new)].value = float(run_avgs["total_avg"])

    print("")
    print("Experiments Database Excel file has been updated successfully.")

    test_date = '{dt.month}/{dt.day}/{dt.year}'.format(dt=datetime.now())
    data['B{}'.format(new)].value = test_date

    previous_runs = []
    for i in range(3, hr+1):
        if data.cell(row=i, column=1).value is not None:
            previous_runs.append(data.cell(row=i, column=1).value)
    previous_run = max(previous_runs)
    this_run = previous_run + 1
    data['A{}'.format(new)].value = this_run

    for i in range(1, 23):
        data.cell(row=new, column=i).alignment = Alignment(horizontal='center')

    wb_exp.save(filename=wb_path)


def main():

    user_dir = get_users_main_directory()
    check_lp_power_data_folder_exists(user_dir)
    check_database_file_exists(user_dir)
    trial = {}
    trial = prompt_setup(trial, user_dir)

    # stopgap for pre-class variables
    walk_speed = trial["Speed [MPH]"]
    walk_style = trial["Walk Style [Normal/LP]"]
    pack_number = trial["Pack Number [XXX]"]

    # MAKE FOLDERS FOR TRIAL BASED ON PREVIOUSLY-ENTERED INPUTS
    make_pack_folder(pack_number, walk_speed, walk_style, user_dir)

    print('.....')
    print("..........Launching.....")

    # SET UP BASIC AUTO-IT SETTINGS TO USE WINDOW-RELATIVE COORDINATES
    autoit.auto_it_set_option('MouseCoordMode', 0)
    autoit.auto_it_set_option('SendKeyDelay', 10)

    # OPEN PICOSCOPE USING OUR DEFAULT WINDOW LAYOUT WITH TWO PANELS AND POWER UP TOP
    default_pico = os.path.join(user_dir, "electrical_power.psdata")
    click_on_file(default_pico)

    # PREPARE TO RUN THE RECORD DATA MACRO
    autoit.win_wait('PicoScope 6 - [electrical_power.psdata]')
    autoit.win_activate('PicoScope 6 - [electrical_power.psdata]')
    time.sleep(0.5)
    autoit.send('{ALT}{TAB}{TAB}{TAB}{TAB}{ENTER}{END}{UP}{ENTER}')
    autoit.win_wait('Macro Recorder')
    autoit.control_click('Macro Recorder', "[Name:_buttonImport]")
    time.sleep(1)
    macro_text = os.path.abspath(os.path.join(user_dir, "record_20sec.psmacro"))
    autoit.clip_put(macro_text)
    autoit.win_wait("Open")
    autoit.send('^V', 0)
    autoit.send('{ENTER}')
    msgbox("Ready to go?")

    # RUN THE MACRO
    autoit.win_activate('Macro Recorder')
    autoit.win_wait('Macro Recorder')
    autoit.control_click('Macro Recorder', "[Name:_buttonExecute]")
    time.sleep(22)
    autoit.win_activate('Macro Recorder')
    autoit.send('{ESCAPE}')

    # BEGIN SAVING THE DATA
    i = datetime.now()
    save_text = os.path.abspath(os.path.join(user_dir, 'LP Power Data', "{}".format("LP-{0}".format(pack_number)), "{0}MPH-{1} Walk".format(walk_speed, walk_style),
                             "{0}MPH-{1} Walk_".format(walk_speed, walk_style) + i.strftime('%Y-%m-%d %Hh%Mm%Ss')+ ".psdata"))
    autoit.clip_put(save_text)

    # SAVE ALL CAPTURED WAVEFORMS IN PICO FORMAT
    autoit.win_activate('PicoScope 6')
    autoit.win_wait("PicoScope 6")
    autoit.send('{ALT}{f}{a}')
    time.sleep(.5)
    autoit.win_wait("Save As")
    autoit.control_send('Save As', "[Class:Edit;INSTANCE:1]", '^V', 0)
    autoit.send('{ENTER}')

    # BEGIN SAVE SEQUENCE FOR CSV FORMAT
    autoit.win_activate('PicoScope 6')
    autoit.win_wait("PicoScope 6")
    autoit.mouse_click("primary", 433, 67, 2, 1)
    # MUST CLICK ON SECTION OF WINDOW CONTAINING POWER TRACE OR ALL AVERAGES CALCULATED WILL BE MESSED UP
    autoit.win_activate('PicoScope 6')
    autoit.win_wait("PicoScope 6")
    autoit.mouse_click("primary", 230, 260, 1, 0)
    # autoit.control_click('PicoScope 6', "[CLASS:WindowsForms10.Window.8.app.0.1114f81_r13_ad1; INSTANCE:55]")
    autoit.send('{ALT}{f}{a}')
    time.sleep(.5)
    autoit.win_wait("Save As")
    autoit.control_send('Save As', "[Class:Edit;INSTANCE:1]", '^V', 0)

    # SAVE ALL CAPTURED 5SEC INTERVALS
    autoit.control_click('Save As', "[CLASS:ComboBox;INSTANCE:3]")
    autoit.send("{HOME}{DOWN}{DOWN}{ENTER}")
    # autoit.control_click("Save As", "[NAME:_currentBufferRadio]")
    time.sleep(0.5)
    autoit.send('{ENTER}')
    time.sleep(4.5)
    autoit.process_close('PicoScope.exe')
    # BOTH SAVE OPERATIONS NOW COMPLETE. THERE IS A CSV FOR EACH 5SEC INTERVAL IN A NEW SUBDIRECTORY OF THIS TRIAL.

    print("")
    print("Sit tight, I'm still doing some thinking....")

    # UPDATE ALL CSV FILES TO BE XLSX AND CALCULATE THEIR 5SEC AVERAGES
    run_avgs = recursively_repair_all_csv_files(save_text, user_dir, walk_speed)

    # ADD ALL INFORMATION COLLECTED TO THE EXPERIMENTS SPREADSHEET
    patch_experiments_database(trial, run_avgs, user_dir)
    print("")

    # WRITE THE METADATA SETTINGS TO A FILE FOR FUTURE RECORD
    with open('{}'.format(save_text.replace(".psdata", ".LPmeta")), 'w') as outfile:
        json.dump(trial, outfile, indent=4, separators=(',', ': '))
    print("")
    print("LP Metadata file successfully saved for the future with all info the state of the pack as you have just tested it.")

    print("")
    # print_average_power(dest_filename)
    print(".................................. Testing Complete.")


if __name__ == '__main__':
    main()
